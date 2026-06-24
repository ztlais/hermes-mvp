import openpyxl

path = "/root/hermes-mvp/projets_toiture.xlsx"
wb = openpyxl.load_workbook(path)
ws = wb.active

print("Colonne A avant correction :")
for row in range(1, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    print(f"  L{row}: {val}")

# Replace "Pr1 – Genouillac" → "Pr1" etc. Also handle header row if needed
for row in range(1, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and isinstance(val, str) and val.startswith("Pr") and "–" in val:
        # Keep only the PrX part
        new_val = val.split("–")[0].strip()
        ws.cell(row=row, column=1).value = new_val
        print(f"  → L{row}: '{val}' → '{new_val}'")

wb.save(path)
print("\n✅ Saved!")
